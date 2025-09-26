import os
import pandas as pd
import matplotlib.pyplot as plt
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# -----------------------------
# Database connection
# -----------------------------
# Database configuration
DB_CONFIG = {
    "dbname": "olist_db",   # ✅ correct name from pgAdmin
    "user": "postgres",
    "password": "99113344",
    "host": "localhost",
    "port": 5432,
}



# -----------------------------
# Helpers
# -----------------------------
def get_connection():
    return psycopg2.connect(**DB_CONFIG)

def fetch_dataframe(query):
    with get_connection() as conn:
        return pd.read_sql_query(query, conn)

def save_chart(df, chart_type, filename, title, xlabel=None, ylabel=None, x=None, y=None):
    plt.figure(figsize=(8, 6))
    if chart_type == "pie":
        df.set_index(x).plot.pie(y=y, autopct="%1.1f%%", legend=False)
    elif chart_type == "bar":
        df.plot.bar(x=x, y=y, legend=False)
    elif chart_type == "barh":
        df.plot.barh(x=x, y=y, legend=False)
    elif chart_type == "line":
        df.plot(x=x, y=y, legend=False)
    elif chart_type == "hist":
        df[y].plot.hist(bins=20)
    elif chart_type == "scatter":
        # convert timedelta column (if exists) into numeric days
        if pd.api.types.is_timedelta64_dtype(df[x]):
            df[x] = df[x].dt.days
        df.plot.scatter(x=x, y=y)

    plt.title(title)
    if xlabel: plt.xlabel(xlabel)
    if ylabel: plt.ylabel(ylabel)
    plt.tight_layout()

    os.makedirs("charts", exist_ok=True)
    path = f"charts/{filename}.png"
    plt.savefig(path)
    plt.close()
    print(f"✅ Saved {chart_type} chart as {path} (rows: {len(df)}) → {title}")

def export_to_excel(dfs, filename):
    os.makedirs("exports", exist_ok=True)
    path = f"exports/{filename}"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    # Apply formatting
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Freeze header
        ws.freeze_panes = "B2"

        # Filters
        ws.auto_filter.ref = ws.dimensions

        # Gradient fill (first numeric column only for demo)
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all(isinstance(c.value, (int, float)) for c in col if c.value is not None):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(path)
    print(f"📊 Created Excel file {filename}, {len(dfs)} sheets")

# -----------------------------
# Main
# -----------------------------
def main():
    # Queries (with JOINs)
    queries = {
        "pie": """
            SELECT p.payment_type, COUNT(*) AS order_count
            FROM payments p
            JOIN orders o ON p.order_id = o.order_id
            GROUP BY p.payment_type;
        """,
        "bar": """
            SELECT pc.product_category_name AS category, COUNT(*) AS product_count
            FROM products pr
            JOIN product_category_name_translation pc
            ON pr.product_category_name = pc.product_category_name
            GROUP BY pc.product_category_name
            ORDER BY product_count DESC
            LIMIT 10;
        """,
        "barh": """
            SELECT s.seller_id, SUM(oi.price) AS total_sales
            FROM sellers s
            JOIN order_items oi ON s.seller_id = oi.seller_id
            GROUP BY s.seller_id
            ORDER BY total_sales DESC
            LIMIT 10;
        """,
        "line": """
            SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month, COUNT(*) AS total_orders
            FROM orders o
            GROUP BY month
            ORDER BY month;
        """,
        "hist": """
            SELECT price FROM order_items;
        """,
        "scatter": """
            SELECT (o.order_delivered_customer_date - o.order_purchase_timestamp) AS shipping_days,
                   oi.price
            FROM orders o
            JOIN order_items oi ON o.order_id = oi.order_id
            WHERE o.order_delivered_customer_date IS NOT NULL;
        """
    }

    # Run queries + charts
    df1 = fetch_dataframe(queries["pie"])
    save_chart(df1, "pie", "pie_payment_type", "Distribution of Orders by Payment Type", x="payment_type", y="order_count")

    df2 = fetch_dataframe(queries["bar"])
    save_chart(df2, "bar", "bar_products_category", "Top 10 Product Categories by Count", x="category", y="product_count")

    df3 = fetch_dataframe(queries["barh"])
    save_chart(df3, "barh", "barh_seller_sales", "Top 10 Sellers by Sales", x="seller_id", y="total_sales")

    df4 = fetch_dataframe(queries["line"])
    save_chart(df4, "line", "line_orders_monthly", "Monthly Orders Over Time", x="month", y="total_orders")

    df5 = fetch_dataframe(queries["hist"])
    save_chart(df5, "hist", "hist_product_prices", "Distribution of Product Prices", y="price")

    df6 = fetch_dataframe(queries["scatter"])
    save_chart(df6, "scatter", "scatter_shipping_vs_price", "Shipping Days vs Price", x="shipping_days", y="price")

    # Export to Excel
    export_to_excel({
        "Payments": df1,
        "Categories": df2,
        "Sellers": df3,
        "Orders": df4,
        "Prices": df5,
        "Shipping_vs_Price": df6
    }, "analytics_results.xlsx")

if __name__ == "__main__":
    main()
