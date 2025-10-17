import os
import pandas as pd
import matplotlib.pyplot as plt
import psycopg2
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
import plotly.express as px

# -----------------------------
# Database connection
# -----------------------------
DB_CONFIG = {
    "dbname": "olist_db",
    "user": "postgres",
    "password": "99113344",
    "host": "localhost",
    "port": 5432,
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

def fetch_dataframe(query):
    with get_connection() as conn:
        return pd.read_sql_query(query, conn)

# -----------------------------
# Chart saving function
# -----------------------------
def save_chart(df, chart_type, filename, title, xlabel=None, ylabel=None, x=None, y=None, log=False):
    if df.empty:
        print(f"⚠️ Skipping {filename} (no data)")
        return

    plt.figure(figsize=(10, 6))
    plt.title(title, fontsize=14, fontweight="bold")

    # --- PIE CHART ---
    if chart_type == "pie":
        df = df[df[x] != "not_defined"]  # remove invalid data
        df.set_index(x).plot.pie(
            y=y, autopct="%1.1f%%", legend=False, colors=plt.cm.Set3.colors
        )

    # --- BARH CHART (Top Sellers) ---
    elif chart_type == "barh":
        # Keep only top 15 sellers for clarity
        df = df.sort_values(y, ascending=True).tail(15)
        bars = plt.barh(
            df[x],
            df[y],
            color=plt.cm.tab20.colors[:len(df)]
        )
        if log:
            plt.xscale("log")
        plt.xlabel(xlabel or x)
        plt.ylabel(ylabel or y)

        # Add labels next to bars
        for bar in bars:
            plt.text(
                bar.get_width() * 1.02,
                bar.get_y() + bar.get_height() / 2,
                f"${bar.get_width():,.0f}",
                va="center",
                fontsize=9,
            )

    # --- OTHER CHART TYPES ---
    elif chart_type == "bar":
        df.plot.bar(x=x, y=y, legend=False)
    elif chart_type == "line":
        df.plot(x=x, y=y, legend=False, marker='o', color='darkcyan')
    elif chart_type == "hist":
        plt.hist(df[y], bins=50, color="skyblue", edgecolor="black", log=True)
    elif chart_type == "scatter":
        if pd.api.types.is_timedelta64_dtype(df[x]):
            df[x] = df[x].dt.days
        plt.scatter(df[x], df[y], alpha=0.4, color="teal")

    plt.grid(True, linestyle="--", alpha=0.6)
    plt.tight_layout()

    os.makedirs("charts", exist_ok=True)
    path = f"charts/{filename}.png"
    plt.savefig(path, dpi=300)
    plt.close()
    print(f"✅ Saved chart: {path} ({len(df)} rows)")

# -----------------------------
# Excel export (with color scale)
# -----------------------------
def export_to_excel(dfs, filename):
    os.makedirs("exports", exist_ok=True)
    path = f"exports/{filename}"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        # Add color scale formatting for numeric columns
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all(isinstance(c.value, (int, float)) for c in col if c.value is not None):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(path)
    print(f"📊 Excel exported successfully → {filename}")

# -----------------------------
# Main script
# -----------------------------
def main():
    queries = {
        "pie": """
            SELECT p.payment_type, COUNT(*) AS order_count
            FROM payments p
            JOIN orders o ON p.order_id = o.order_id
            GROUP BY p.payment_type;
        """,
        "barh": """
            SELECT s.seller_id, SUM(oi.price) AS total_sales
            FROM sellers s
            JOIN order_items oi ON s.seller_id = oi.seller_id
            GROUP BY s.seller_id
            ORDER BY total_sales DESC;
        """,
        "line": """
            SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month, COUNT(*) AS total_orders
            FROM orders o
            GROUP BY month
            ORDER BY month;
        """,
        "hist": "SELECT price FROM order_items;",
        "scatter": """
            SELECT (o.order_delivered_customer_date - o.order_purchase_timestamp) AS shipping_days,
                   oi.price
            FROM orders o
            JOIN order_items oi ON o.order_id = oi.order_id
            WHERE o.order_delivered_customer_date IS NOT NULL;
        """
    }

    # --- PIE ---
    df1 = fetch_dataframe(queries["pie"])
    save_chart(df1, "pie", "pie_payment_type", "Distribution of Orders by Payment Type", x="payment_type", y="order_count")

    # --- BARH ---
    df2 = fetch_dataframe(queries["barh"])
    save_chart(df2, "barh", "barh_seller_sales", "Top 15 Sellers by Sales (Log Scale)", x="seller_id", y="total_sales", log=True)

    # --- LINE ---
    df3 = fetch_dataframe(queries["line"])
    save_chart(df3, "line", "line_orders_monthly", "Monthly Orders Over Time", x="month", y="total_orders")

    # --- HIST ---
    df4 = fetch_dataframe(queries["hist"])
    q95_price = df4["price"].quantile(0.95)
    df4_filtered = df4[df4["price"] <= q95_price]
    save_chart(df4_filtered, "hist", "hist_product_prices", "Distribution of Product Prices (Log Scale)", y="price")

    # --- SCATTER ---
    df5 = fetch_dataframe(queries["scatter"])
    q95_price2 = df5["price"].quantile(0.95)
    df5_filtered = df5[df5["price"] <= q95_price2]
    save_chart(df5_filtered, "scatter", "scatter_shipping_vs_price", "Shipping Days vs Price", x="shipping_days", y="price")

    # --- EXCEL EXPORT ---
    export_to_excel({
        "Payments": df1,
        "Sellers": df2.head(50),
        "Orders": df3,
        "Prices": df4_filtered,
        "Shipping_vs_Price": df5_filtered
    }, "analytics_results.xlsx")

    # --- INTERACTIVE CHART ---
    df_time = df3.copy()
    df_time["year"] = df_time["month"].dt.year
    df_time["month_name"] = df_time["month"].dt.strftime("%b %Y")
    fig = px.bar(
        df_time,
        x="month_name",
        y="total_orders",
        color="year",
        animation_frame="year",
        title="📈 Monthly Orders Over Time (Interactive)"
    )
    fig.show()

if __name__ == "__main__":
    main()
